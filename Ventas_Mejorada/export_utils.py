# ==================== MÓDULO DE EXPORTACIÓN v1.0 ====================
# Exportación a Excel (.xlsx) y PDF para Locatel AIS
# Requiere: openpyxl, reportlab
# Instalar: pip install openpyxl reportlab

import io
import base64
from datetime import datetime
import pandas as pd
import streamlit as st

# ── Paleta corporativa ──────────────────────────────────────────────
COLOR_PRIMARY    = "1A56DB"
COLOR_PRIMARY_DK = "1341B3"
COLOR_SUCCESS    = "16A34A"
COLOR_WARNING    = "D97706"
COLOR_DANGER     = "DC2626"
COLOR_HEADER_BG  = "0F172A"   # sidebar dark
COLOR_ROW_ALT    = "F8FAFC"
COLOR_BORDER     = "E2E8F0"
COLOR_WHITE      = "FFFFFF"


# ════════════════════════════════════════════════════════════════════
#  EXCEL
# ════════════════════════════════════════════════════════════════════
def exportar_excel(df: pd.DataFrame, titulo: str, subtitulo: str = "", hoja: str = "Reporte") -> bytes:
    """
    Genera un archivo Excel profesional con el DataFrame recibido.
    Devuelve bytes listos para st.download_button.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        st.error("📦 Instala openpyxl: `pip install openpyxl`")
        return b""

    wb = Workbook()
    ws = wb.active
    ws.title = hoja

    # ── Helpers de estilo ──
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border(color=COLOR_BORDER):
        side = Side(style="thin", color=color)
        return Border(left=side, right=side, top=side, bottom=side)

    def font(bold=False, color=COLOR_HEADER_BG, size=11, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ── Fila 1: Título principal ──
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    c = ws["A1"]
    c.value        = titulo.upper()
    c.font         = Font(bold=True, color=COLOR_WHITE, size=14, name="Calibri")
    c.fill         = fill(COLOR_HEADER_BG)
    c.alignment    = align("center")
    ws.row_dimensions[1].height = 36

    # ── Fila 2: Subtítulo / fecha ──
    fecha_str = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  {subtitulo}"
    ws.merge_cells(f"A2:{get_column_letter(len(df.columns))}2")
    c = ws["A2"]
    c.value     = fecha_str
    c.font      = Font(italic=True, color="94A3B8", size=10, name="Calibri")
    c.fill      = fill("1E293B")
    c.alignment = align("center")
    ws.row_dimensions[2].height = 22

    # ── Fila 3: vacía (separador) ──
    ws.row_dimensions[3].height = 6

    # ── Fila 4: Encabezados ──
    header_row = 4
    for col_idx, col_name in enumerate(df.columns, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=str(col_name).upper())
        c.font      = Font(bold=True, color=COLOR_WHITE, size=10, name="Calibri")
        c.fill      = fill(COLOR_PRIMARY)
        c.alignment = align("center")
        c.border    = border(COLOR_WHITE)
    ws.row_dimensions[header_row].height = 24

    # ── Filas de datos ──
    for row_idx, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        is_alt = (row_idx - header_row) % 2 == 0
        bg = COLOR_ROW_ALT if is_alt else COLOR_WHITE
        for col_idx, value in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.fill      = fill(bg)
            c.border    = border()
            c.font      = font(size=10)
            # Alineación según tipo
            if isinstance(value, (int, float)):
                c.alignment = align("right")
                if isinstance(value, float):
                    c.number_format = "#,##0.00"
                else:
                    c.number_format = "#,##0"
            else:
                c.alignment = align("left")
        ws.row_dimensions[row_idx].height = 20

    # ── Ancho de columnas automático ──
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *(len(str(v)) for v in df.iloc[:, col_idx - 1].astype(str))
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # ── Freeze panes ──
    ws.freeze_panes = f"A{header_row + 1}"

    # ── Auto-filter ──
    last_col = get_column_letter(len(df.columns))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(df)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════════════
#  PDF
# ════════════════════════════════════════════════════════════════════
def exportar_pdf(df: pd.DataFrame, titulo: str, subtitulo: str = "") -> bytes:
    """
    Genera un PDF profesional con el DataFrame recibido.
    Devuelve bytes listos para st.download_button.
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        st.error("📦 Instala reportlab: `pip install reportlab`")
        return b""

    buf = io.BytesIO()

    # Orientación según cantidad de columnas
    page_size = landscape(A4) if len(df.columns) > 6 else A4
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    # ── Colores ──
    C_PRIMARY = colors.HexColor(f"#{COLOR_PRIMARY}")
    C_DARK    = colors.HexColor(f"#{COLOR_HEADER_BG}")
    C_ALT     = colors.HexColor(f"#{COLOR_ROW_ALT}")
    C_BORDER  = colors.HexColor(f"#{COLOR_BORDER}")
    C_SUCCESS = colors.HexColor(f"#{COLOR_SUCCESS}")
    C_WHITE   = colors.white

    # ── Estilos de texto ──
    styles = getSampleStyleSheet()
    style_titulo = ParagraphStyle(
        "titulo", parent=styles["Title"],
        fontSize=16, fontName="Helvetica-Bold",
        textColor=C_DARK, alignment=TA_CENTER, spaceAfter=4
    )
    style_sub = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica-Oblique",
        textColor=colors.HexColor("#94A3B8"),
        alignment=TA_CENTER, spaceAfter=12
    )
    style_footer = ParagraphStyle(
        "footer", parent=styles["Normal"],
        fontSize=8, fontName="Helvetica",
        textColor=colors.HexColor("#94A3B8"),
        alignment=TA_CENTER
    )

    elements = []

    # ── Encabezado ──
    elements.append(Paragraph(titulo, style_titulo))
    fecha_str = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    if subtitulo:
        fecha_str = f"{subtitulo}   ·   {fecha_str}"
    elements.append(Paragraph(fecha_str, style_sub))
    elements.append(HRFlowable(width="100%", thickness=2, color=C_PRIMARY, spaceAfter=12))

    # ── Tabla ──
    page_w = page_size[0] - 3*cm
    n_cols = len(df.columns)
    col_w  = [page_w / n_cols] * n_cols

    # Encabezados
    headers = [str(c).upper() for c in df.columns]
    data    = [headers]

    for _, row in df.iterrows():
        data.append([str(v) if v is not None else "" for v in row])

    table = Table(data, colWidths=col_w, repeatRows=1)

    # Estilo de la tabla
    ts = TableStyle([
        # Encabezado
        ("BACKGROUND",  (0, 0), (-1, 0), C_PRIMARY),
        ("TEXTCOLOR",   (0, 0), (-1, 0), C_WHITE),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 9),
        ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
        ("TOPPADDING",  (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        # Datos
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("ALIGN",       (0, 1), (-1, -1), "LEFT"),
        ("TOPPADDING",  (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        # Bordes
        ("GRID",        (0, 0), (-1, -1), 0.5, C_BORDER),
        ("LINEBELOW",   (0, 0), (-1, 0),  1.5, C_WHITE),
        # Filas alternas
        *[
            ("BACKGROUND", (0, i), (-1, i), C_ALT)
            for i in range(2, len(data), 2)
        ],
    ])
    table.setStyle(ts)
    elements.append(table)

    # ── Resumen al pie ──
    elements.append(Spacer(1, 0.4*cm))
    elements.append(HRFlowable(width="100%", thickness=1, color=C_BORDER))
    elements.append(Spacer(1, 0.2*cm))
    total_rows = len(df)
    elements.append(
        Paragraph(
            f"Total de registros: <b>{total_rows}</b>   ·   Locatel AIS © {datetime.now().year}",
            style_footer
        )
    )

    doc.build(elements)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════════════
#  HELPERS DE STREAMLIT
# ════════════════════════════════════════════════════════════════════
def boton_descarga_excel(
    df: pd.DataFrame,
    titulo: str,
    subtitulo: str = "",
    nombre_archivo: str = "reporte",
    label: str = "⬇️ Descargar Excel",
    key: str = None,
):
    """Renderiza un botón de descarga Excel de Streamlit."""
    data = exportar_excel(df, titulo, subtitulo)
    if data:
        fname = f"{nombre_archivo}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            label=label,
            data=data,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=key or f"dl_excel_{nombre_archivo}",
        )


def boton_descarga_pdf(
    df: pd.DataFrame,
    titulo: str,
    subtitulo: str = "",
    nombre_archivo: str = "reporte",
    label: str = "⬇️ Descargar PDF",
    key: str = None,
):
    """Renderiza un botón de descarga PDF de Streamlit."""
    data = exportar_pdf(df, titulo, subtitulo)
    if data:
        fname = f"{nombre_archivo}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        st.download_button(
            label=label,
            data=data,
            file_name=fname,
            mime="application/pdf",
            key=key or f"dl_pdf_{nombre_archivo}",
        )


def barra_exportacion(
    df: pd.DataFrame,
    titulo: str,
    subtitulo: str = "",
    nombre_archivo: str = "reporte",
    key_prefix: str = "exp",
):
    """
    Renderiza una fila con los dos botones de exportación (Excel + PDF)
    alineados a la derecha. Llamar DESPUÉS de mostrar la tabla.
    """
    if df is None or df.empty:
        return
    _, col_excel, col_pdf = st.columns([6, 1, 1])
    with col_excel:
        boton_descarga_excel(
            df, titulo, subtitulo, nombre_archivo,
            label="📊 Excel",
            key=f"{key_prefix}_excel"
        )
    with col_pdf:
        boton_descarga_pdf(
            df, titulo, subtitulo, nombre_archivo,
            label="📄 PDF",
            key=f"{key_prefix}_pdf"
        )
